from __future__ import annotations
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
import scipy.io.wavfile as scipy_wavfile
from openpyxl import load_workbook
from PyQt6 import QtWidgets, QtCore, QtGui
from PyQt6.QtCore import pyqtSignal, QUrl
from PyQt6.QtMultimedia import QMediaPlayer, QAudioOutput

import matplotlib
matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as Canvas
from matplotlib.figure import Figure
import matplotlib.ticker as ticker

# Configure matplotlib to support Chinese
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial', 'sans-serif']
matplotlib.rcParams['axes.unicode_minus'] = False  # Ensure minus signs are shown correctly


from phonetic_toolbox.services.settings_service import SettingsService
from phonetic_toolbox.services.acoustic_service import PARAMETER_MAPPING
from phonetic_toolbox.services.io.wav import read_wav_float_mono
from phonetic_toolbox.services.io.excel import resolve_parameter_source, load_fastdb_columns, load_fastdb_window, ensure_fast_parameter_db
from phonetic_toolbox.gui.dialogs.parameter_tools_dialog import ParameterHelpDialog as SharedParameterHelpDialog

log = logging.getLogger(__name__)

class ParameterDisplayWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.settings = SettingsService()
        self._current_wav_path: Optional[Path] = None
        self._current_xlsx_path: Optional[Path] = None
        self._param_file_path: Optional[Path] = None
        self._param_file_format: Optional[str] = None
        self._time_index: Optional[np.ndarray] = None
        self._time_row_index: Optional[np.ndarray] = None
        self._textgrid_cols: List[str] = []
        self._numeric_cols: List[str] = []
        self._player = QMediaPlayer()
        self._audio_output = QAudioOutput()
        self._player.setAudioOutput(self._audio_output)
        self._playback_end_ms: Optional[int] = None
        self._pending_play_start_ms: Optional[int] = None
        self._pending_play_end_ms: Optional[int] = None
        self._player.positionChanged.connect(self._on_player_position_changed)
        self._player.mediaStatusChanged.connect(self._on_player_media_status_changed)
        
        self.is_dark = True
        
        self.init_ui()
        self._load_last_dirs()

    def init_ui(self):
        main_layout = QtWidgets.QHBoxLayout(self)
        
        # --- Left: File List ---
        left_layout = QtWidgets.QVBoxLayout()
        
        # WAV Dir
        wav_group = QtWidgets.QGroupBox("WAV目录")
        wav_layout = QtWidgets.QVBoxLayout()
        self.edit_wav_dir = QtWidgets.QLineEdit()
        self.btn_browse_wav = QtWidgets.QPushButton("浏览...")
        h1 = QtWidgets.QHBoxLayout()
        h1.addWidget(self.edit_wav_dir)
        h1.addWidget(self.btn_browse_wav)
        wav_layout.addLayout(h1)
        wav_group.setLayout(wav_layout)
        
        # XLSX Dir
        xlsx_group = QtWidgets.QGroupBox("XLSX目录")
        xlsx_layout = QtWidgets.QVBoxLayout()
        self.edit_xlsx_dir = QtWidgets.QLineEdit()
        self.btn_browse_xlsx = QtWidgets.QPushButton("浏览...")
        h2 = QtWidgets.QHBoxLayout()
        h2.addWidget(self.edit_xlsx_dir)
        h2.addWidget(self.btn_browse_xlsx)
        xlsx_layout.addLayout(h2)
        xlsx_group.setLayout(xlsx_layout)
        
        # File List
        self.list_files = QtWidgets.QListWidget()
        self.edit_filter_files = QtWidgets.QLineEdit()
        self.edit_filter_files.setPlaceholderText("筛选音频 (输入字符串)")
        
        left_layout.addWidget(wav_group)
        left_layout.addWidget(xlsx_group)
        left_layout.addWidget(self.edit_filter_files)
        left_layout.addWidget(self.list_files)
        
        # Controls
        self.btn_refresh = QtWidgets.QPushButton("刷新")
        left_layout.addWidget(self.btn_refresh)
        
        # Wrapper for left side width
        left_widget = QtWidgets.QWidget()
        left_widget.setLayout(left_layout)
        left_widget.setMaximumWidth(300)
        
        # --- Middle: Parameter List ---
        mid_layout = QtWidgets.QVBoxLayout()
        mid_layout.addWidget(QtWidgets.QLabel("参数列表"))
        self.edit_filter_params = QtWidgets.QLineEdit()
        self.edit_filter_params.setPlaceholderText("筛选参数")
        self.list_params = QtWidgets.QListWidget()
        self.list_params.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        
        # Checkboxes for filtering
        self.chk_reaper = QtWidgets.QCheckBox("reaper")
        self.chk_reaper.setChecked(True)
        self.chk_correction = QtWidgets.QCheckBox("correction")
        self.chk_correction.setChecked(False)
        
        mid_layout.addWidget(self.edit_filter_params)
        mid_layout.addWidget(self.list_params)
        mid_layout.addWidget(self.chk_reaper)
        mid_layout.addWidget(self.chk_correction)
        
        mid_widget = QtWidgets.QWidget()
        mid_widget.setLayout(mid_layout)
        mid_widget.setMaximumWidth(200)
        
        # --- Right: Plot Area ---
        right_layout = QtWidgets.QVBoxLayout()
        top_bar = QtWidgets.QWidget()
        top_bar.setFixedHeight(44)
        top_bar_layout = QtWidgets.QHBoxLayout(top_bar)
        top_bar_layout.setContentsMargins(0, 0, 0, 0)
        top_bar_layout.setSpacing(8)
        self.lbl_window_range = QtWidgets.QLabel("区间: -")
        self.lbl_window_range.setFixedHeight(20)
        self.slider_position = QtWidgets.QSlider(QtCore.Qt.Orientation.Horizontal)
        self.slider_position.setFixedHeight(20)
        self.slider_position.setMinimum(0)
        self.slider_position.setMaximum(0)
        self.slider_position.setValue(0)
        self.slider_position.setTracking(False)
        self.slider_position.setEnabled(False)
        top_bar_layout.addWidget(self.lbl_window_range)
        top_bar_layout.addWidget(self.slider_position, 1)
        right_layout.addWidget(top_bar)
        
        self.figure = Figure()
        self.canvas = Canvas(self.figure)
        self.ax_wave = self.figure.add_subplot(211)
        self.ax_param = self.figure.add_subplot(212, sharex=self.ax_wave)
        
        right_layout.addWidget(self.canvas)
        
        # Play controls
        play_layout = QtWidgets.QHBoxLayout()
        self.btn_play = QtWidgets.QPushButton("播放")
        self.btn_save_img = QtWidgets.QPushButton("保存图片")
        self.btn_param_help = QtWidgets.QPushButton("参数说明")
        self.btn_help_doc = QtWidgets.QPushButton("帮助")
        self.btn_help_doc.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                border: 1px solid #1f7a33;
                border-radius: 4px;
                padding: 4px 10px;
            }
            QPushButton:hover {
                background-color: #34c759;
            }
            QPushButton:pressed {
                background-color: #1f7a33;
            }
        """)
        self.btn_close = QtWidgets.QPushButton("关闭")
        
        play_layout.addWidget(self.btn_play)
        play_layout.addWidget(self.btn_save_img)
        play_layout.addWidget(self.btn_param_help)
        play_layout.addWidget(self.btn_help_doc)
        play_layout.addStretch()
        play_layout.addWidget(self.btn_close)
        
        right_layout.addLayout(play_layout)
        
        # Add to main
        main_layout.addWidget(left_widget)
        main_layout.addWidget(mid_widget)
        main_layout.addLayout(right_layout)
        
        # Connections
        self.btn_browse_wav.clicked.connect(self._browse_wav)
        self.btn_browse_xlsx.clicked.connect(self._browse_xlsx)
        self.edit_wav_dir.editingFinished.connect(self._on_wav_dir_edited)
        self.btn_refresh.clicked.connect(self._refresh)
        self.edit_filter_files.textChanged.connect(self._filter_files)
        self.edit_filter_params.textChanged.connect(self._update_param_visibility)
        self.chk_reaper.stateChanged.connect(self._update_param_visibility)
        self.chk_correction.stateChanged.connect(self._update_param_visibility)
        self.list_files.itemSelectionChanged.connect(self._on_file_selected)
        self.list_params.itemSelectionChanged.connect(self._plot)
        self.btn_play.clicked.connect(self._play_audio)
        self.btn_save_img.clicked.connect(self._save_image)
        self.btn_param_help.clicked.connect(self._show_param_help)
        self.btn_help_doc.clicked.connect(self._open_help_doc)
        self.btn_close.clicked.connect(self.close)
        self.slider_position.valueChanged.connect(self._on_position_slider_changed)
        
        # Interaction variables
        self._wav_data: Optional[np.ndarray] = None
        self._wav_dtype = None
        self._wav_total_samples: int = 0
        self._fs: int = 16000
        self._press_x: Optional[float] = None
        self._pan_active = False
        self._max_window_sec = 10.0
        self._view_start_sec = 0.0
        self._view_width_sec = self._max_window_sec
        self._slider_updating = False

        self.canvas.mpl_connect("scroll_event", self._on_scroll)
        self.canvas.mpl_connect("button_press_event", self._on_press)
        self.canvas.mpl_connect("button_release_event", self._on_release)
        self.canvas.mpl_connect("motion_notify_event", self._on_motion)

    def _load_last_dirs(self):
        wav_dir = self.settings.get("pd_wav_dir")
        xlsx_dir = self.settings.get("pd_xlsx_dir")
        if wav_dir:
            self.edit_wav_dir.setText(wav_dir)
        if xlsx_dir:
            self.edit_xlsx_dir.setText(xlsx_dir)
        
        if wav_dir and Path(wav_dir).exists():
            self._refresh()

    def _browse_wav(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "选择WAV目录", self.edit_wav_dir.text())
        if d:
            self._set_wav_dir(d, sync_xlsx=True)
            self._refresh()

    def _browse_xlsx(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "选择XLSX目录", self.edit_xlsx_dir.text())
        if d:
            self.edit_xlsx_dir.setText(d)
            self.settings.set("pd_xlsx_dir", d)
            self._refresh()

    def _on_wav_dir_edited(self):
        d = self.edit_wav_dir.text().strip()
        if not d:
            return
        self._set_wav_dir(d, sync_xlsx=True)
        self._refresh()

    def _set_wav_dir(self, directory: str, sync_xlsx: bool):
        self.edit_wav_dir.setText(directory)
        self.settings.set("pd_wav_dir", directory)
        if sync_xlsx:
            self.edit_xlsx_dir.setText(directory)
            self.settings.set("pd_xlsx_dir", directory)

    def _refresh(self):
        self.list_files.clear()
        p = Path(self.edit_wav_dir.text())
        if not p.exists():
            return
            
        files = sorted(list(p.glob("*.wav")))
        for f in files:
            self.list_files.addItem(f.name)
            
        self._filter_files(self.edit_filter_files.text())

    def _filter_files(self, text):
        count = self.list_files.count()
        for i in range(count):
            item = self.list_files.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def _update_param_visibility(self):
        text = self.edit_filter_params.text().lower()
        show_reaper = self.chk_reaper.isChecked()
        show_correction = self.chk_correction.isChecked()
        
        count = self.list_params.count()
        for i in range(count):
            item = self.list_params.item(i)
            item_text = item.text()
            item_text_lower = item_text.lower()
            
            # Check text filter
            if text and text not in item_text_lower:
                item.setHidden(True)
                continue
                
            # Check reaper filter
            if not show_reaper and "rf0" in item_text_lower:
                item.setHidden(True)
                continue
                
            # Check correction filter
            if not show_correction and "*" in item_text:
                item.setHidden(True)
                continue
                
            item.setHidden(False)

    def _on_file_selected(self):
        items = self.list_files.selectedItems()
        if not items:
            return
        
        fname = items[0].text()
        wav_path = Path(self.edit_wav_dir.text()) / fname
        stem = Path(fname).stem
        param_base_path = Path(self.edit_xlsx_dir.text()) / f"{stem}.xlsx"
        
        self._current_wav_path = wav_path
        self._current_xlsx_path = param_base_path
        self._wav_data = None
        self._wav_dtype = None
        self._wav_total_samples = 0
        self._fs = 16000
        self._param_file_path = None
        self._param_file_format = None
        self._time_index = None
        self._time_row_index = None
        
        self.list_params.clear()
        self._textgrid_cols = []
        self._numeric_cols = []

        ensured_fast_path = ensure_fast_parameter_db(param_base_path)
        resolved_path, resolved_format = resolve_parameter_source(param_base_path)
        if resolved_format == "fastdb" and (ensured_fast_path is None or resolved_path != ensured_fast_path):
            xlsx_fallback = param_base_path
            csv_fallback = param_base_path.with_suffix(".csv")
            if xlsx_fallback.exists():
                resolved_path, resolved_format = xlsx_fallback, "xlsx"
            elif csv_fallback.exists():
                resolved_path, resolved_format = csv_fallback, "csv"
            else:
                resolved_path, resolved_format = None, None
        if resolved_path is not None and resolved_format is not None:
            self._param_file_path = resolved_path
            self._param_file_format = resolved_format

        self._load_param_columns_meta()

        if self._current_wav_path and self._current_wav_path.exists():
            try:
                fs, data = scipy_wavfile.read(str(self._current_wav_path), mmap=True)
                self._fs = fs
                self._wav_data = data
                self._wav_dtype = data.dtype
                self._wav_total_samples = int(data.shape[0]) if hasattr(data, "shape") and len(data.shape) > 0 else 0
            except Exception as e:
                log.warning(f"Waveform memmap load failed, fallback to full load: {e}")
                try:
                    fs, y = read_wav_float_mono(self._current_wav_path)
                    self._fs = fs
                    self._wav_data = y
                    self._wav_dtype = y.dtype
                    self._wav_total_samples = len(y)
                except Exception as e2:
                    log.error(f"Waveform load error: {e2}")
                    self._wav_data = None
                    self._wav_dtype = None
                    self._wav_total_samples = 0

        self._configure_view_for_current_audio()
        self._plot()

    def _load_param_columns_meta(self):
        if self._param_file_path is None:
            return
        try:
            if self._param_file_format == "fastdb":
                cols = load_fastdb_columns(self._param_file_path)
            elif self._param_file_format == "csv":
                header_df = pd.read_csv(self._param_file_path, nrows=0)
                cols = list(header_df.columns)
            else:
                wb = load_workbook(filename=str(self._param_file_path), read_only=True, data_only=True)
                ws = wb.active
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                cols = [str(c) for c in first_row if c is not None] if first_row is not None else []
                wb.close()
            for c in cols:
                if c == "Time_s":
                    continue
                if c.startswith("text_"):
                    self._textgrid_cols.append(c)
                else:
                    self._numeric_cols.append(c)

            for c in self._textgrid_cols:
                display_name = f"{c[5:]} (TextGrid)"
                item = QtWidgets.QListWidgetItem(display_name)
                item.setData(QtCore.Qt.ItemDataRole.UserRole, c)
                self.list_params.addItem(item)

            for c in self._numeric_cols:
                item = QtWidgets.QListWidgetItem(c)
                item.setData(QtCore.Qt.ItemDataRole.UserRole, c)
                self.list_params.addItem(item)
        except Exception as e:
            log.error(f"Error loading parameter metadata: {e}")

    def _ensure_time_index(self) -> bool:
        if self._time_index is not None:
            return True
        if self._param_file_path is None:
            return False
        if self._param_file_format == "fastdb":
            return False
        try:
            if self._param_file_format == "csv":
                time_df = pd.read_csv(self._param_file_path, usecols=["Time_s"])
                raw_times = pd.to_numeric(time_df["Time_s"], errors='coerce')
                valid_mask = np.isfinite(raw_times.to_numpy(dtype=float))
                times = raw_times.to_numpy(dtype=float)[valid_mask]
                row_idx = np.flatnonzero(valid_mask).astype(int)
            else:
                wb = load_workbook(filename=str(self._param_file_path), read_only=True, data_only=True)
                ws = wb.active
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if header is None:
                    wb.close()
                    return False
                time_idx = None
                for idx, col_name in enumerate(header):
                    if str(col_name) == "Time_s":
                        time_idx = idx
                        break
                if time_idx is None:
                    wb.close()
                    return False
                times_list = []
                row_idx_list = []
                data_row_index = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if time_idx >= len(row):
                        data_row_index += 1
                        continue
                    v = row[time_idx]
                    if v is None:
                        data_row_index += 1
                        continue
                    times_list.append(v)
                    row_idx_list.append(data_row_index)
                    data_row_index += 1
                wb.close()
                raw_times = pd.to_numeric(pd.Series(times_list), errors='coerce').to_numpy(dtype=float)
                valid_mask = np.isfinite(raw_times)
                times = raw_times[valid_mask]
                row_idx = np.asarray(row_idx_list, dtype=int)[valid_mask]
            if len(times) == 0:
                return False
            self._time_index = times
            self._time_row_index = row_idx
            return True
        except Exception as e:
            log.error(f"Error loading time index: {e}")
            self._time_index = None
            self._time_row_index = None
            return False

    def _load_window_parameter_df(self, start_sec: float, end_sec: float, cols: List[str]) -> Optional[pd.DataFrame]:
        if self._param_file_path is None:
            return None
        if self._param_file_format == "fastdb":
            return load_fastdb_window(self._param_file_path, start_sec, end_sec, cols)
        if not self._ensure_time_index():
            return None
        if self._time_index is None or len(self._time_index) == 0 or self._time_row_index is None or len(self._time_row_index) == 0:
            return None
        if len(self._time_index) >= 2 and np.any(np.diff(self._time_index) < 0):
            return self._load_window_parameter_df_full_scan(start_sec, end_sec, cols)

        needed_cols = ["Time_s"] + [c for c in cols if c != "Time_s"]
        left_idx = int(np.searchsorted(self._time_index, start_sec, side='left'))
        right_idx = int(np.searchsorted(self._time_index, end_sec, side='right'))
        margin = 16
        left_row = int(self._time_row_index[left_idx]) if left_idx < len(self._time_row_index) else int(self._time_row_index[-1])
        right_row = int(self._time_row_index[right_idx - 1]) if right_idx > 0 else int(self._time_row_index[0])
        row_start = max(0, left_row - margin)
        row_end = max(row_start + 1, right_row + margin + 1)
        row_count = max(1, row_end - row_start)
        skiprows = range(1, row_start + 1) if row_start > 0 else None
        try:
            if self._param_file_format == "csv":
                win_df = pd.read_csv(
                    self._param_file_path,
                    usecols=needed_cols,
                    skiprows=skiprows,
                    nrows=row_count
                )
            else:
                win_df = pd.read_excel(
                    self._param_file_path,
                    usecols=needed_cols,
                    skiprows=skiprows,
                    nrows=row_count
                )
            if "Time_s" not in win_df.columns:
                return None
            win_df["Time_s"] = pd.to_numeric(win_df["Time_s"], errors='coerce')
            win_df = win_df[np.isfinite(win_df["Time_s"])]
            win_df = win_df[(win_df["Time_s"] >= start_sec) & (win_df["Time_s"] <= end_sec)]
            if len(win_df) == 0:
                return None
            return win_df.reset_index(drop=True)
        except Exception as e:
            log.error(f"Error loading window parameter data: {e}")
            return None

    def _load_window_parameter_df_full_scan(self, start_sec: float, end_sec: float, cols: List[str]) -> Optional[pd.DataFrame]:
        if self._param_file_path is None:
            return None
        needed_cols = ["Time_s"] + [c for c in cols if c != "Time_s"]
        try:
            if self._param_file_format == "csv":
                full_df = pd.read_csv(self._param_file_path, usecols=needed_cols)
            else:
                full_df = pd.read_excel(self._param_file_path, usecols=needed_cols)
            if "Time_s" not in full_df.columns:
                return None
            full_df["Time_s"] = pd.to_numeric(full_df["Time_s"], errors='coerce')
            full_df = full_df[np.isfinite(full_df["Time_s"])]
            full_df = full_df[(full_df["Time_s"] >= start_sec) & (full_df["Time_s"] <= end_sec)]
            if len(full_df) == 0:
                return None
            return full_df.sort_values("Time_s").reset_index(drop=True)
        except Exception as e:
            log.error(f"Error loading full-scan window data: {e}")
            return None

    def _plot(self):
        self.ax_wave.clear()
        self.ax_param.clear()
        # Handle dual axis clearing?
        if hasattr(self, 'ax_param2'):
            self.ax_param2.remove()
            del self.ax_param2
            
        self._normalize_view_window()
        view_start = self._view_start_sec
        view_end = self._view_start_sec + self._view_width_sec

        if self._wav_data is not None and self._wav_total_samples > 0:
            self.ax_wave.set_ylabel("Amplitude")
            self.ax_wave.set_xlim(view_start, view_end)
            self.ax_param.set_xlim(view_start, view_end)
            self._update_wave_plot()

        selected_items = self.list_params.selectedItems()
        if not selected_items:
            self.ax_wave.set_xlim(view_start, view_end)
            self.ax_param.set_xlim(view_start, view_end)
            self.canvas.draw()
            return

        sel_numeric = []
        sel_textgrid = []
        
        for item in selected_items:
            col = item.data(QtCore.Qt.ItemDataRole.UserRole)
            if col in self._numeric_cols:
                sel_numeric.append(col)
            elif col in self._textgrid_cols:
                sel_textgrid.append(col)

        selected_cols = sel_numeric + sel_textgrid
        window_df = self._load_window_parameter_df(view_start, view_end, selected_cols)
        if window_df is None:
            self.ax_wave.set_xlim(view_start, view_end)
            self.ax_param.set_xlim(view_start, view_end)
            self.set_theme(self.is_dark)
            self.canvas.draw()
            return
        window_df = window_df.sort_values("Time_s").reset_index(drop=True)

        time_vec = pd.to_numeric(window_df["Time_s"], errors='coerce').to_numpy(dtype=float)
        valid_time = np.isfinite(time_vec)
        time_vec = time_vec[valid_time]
        if len(time_vec) == 0:
            self.ax_wave.set_xlim(view_start, view_end)
            self.ax_param.set_xlim(view_start, view_end)
            self.set_theme(self.is_dark)
            self.canvas.draw()
            return
                
        if sel_numeric and len(time_vec) > 0:
            cluster1 = []
            cluster2 = []
            
            means = {}
            for col in sel_numeric:
                if col not in window_df.columns:
                    means[col] = 0
                    continue
                vals = pd.to_numeric(window_df[col], errors='coerce').to_numpy(dtype=float)
                vals = vals[valid_time]
                vals = vals[np.isfinite(vals)]
                if len(vals) == 0:
                    means[col] = 0
                else:
                    means[col] = np.mean(np.abs(vals))
            
            if not means:
                pass
            else:
                # Sort by mean
                sorted_cols = sorted(means.keys(), key=lambda x: means[x])
                
                min_mean = means[sorted_cols[0]]
                max_mean = means[sorted_cols[-1]]
                
                if min_mean > 0 and max_mean / min_mean > 50:
                    for col in sel_numeric:
                        if means[col] > 100:
                            cluster2.append(col)
                        else:
                            cluster1.append(col)
                else:
                    cluster1 = sel_numeric
            
            if cluster1:
                for col in cluster1:
                    if col not in window_df.columns:
                        continue
                    vals = pd.to_numeric(window_df[col], errors='coerce').to_numpy(dtype=float)
                    vals = vals[valid_time]
                    self.ax_param.plot(time_vec, vals, label=col)
                self.ax_param.legend(loc='upper left')
                
            if cluster2:
                self.ax_param2 = self.ax_param.twinx()
                for col in cluster2:
                    if col not in window_df.columns:
                        continue
                    vals = pd.to_numeric(window_df[col], errors='coerce').to_numpy(dtype=float)
                    vals = vals[valid_time]
                    self.ax_param2.plot(time_vec, vals, linestyle='--', label=col)
                self.ax_param2.legend(loc='upper right')
                
                color = 'white' if self.is_dark else 'black'
                self.ax_param2.tick_params(colors=color)
                self.ax_param2.spines['bottom'].set_color(color)
                self.ax_param2.spines['top'].set_color(color)
                self.ax_param2.spines['left'].set_color(color)
                self.ax_param2.spines['right'].set_color(color)

        self.ax_param.set_xlabel("Time (s)")
        self.ax_param.tick_params(labelbottom=True)

        axes_to_draw = [self.ax_wave, self.ax_param]
        line_colors = ['red', 'green', 'orange', 'purple']
        label_color = '#111111' if not self.is_dark else '#f5f5f5'
        
        for i, col in enumerate(sel_textgrid):
            if col not in window_df.columns:
                continue
            vals = window_df[col].astype(str).to_numpy()
            arr = vals[valid_time]
            if len(arr) == 0:
                continue

            shifted = np.roll(arr, 1)
            shifted[0] = object()
            
            changes = (arr != shifted)
            change_indices = np.where(changes)[0]
            
            color = line_colors[i % len(line_colors)]
            y_pos_wave = 0.86 - (i * 0.12)
            segment_ends = np.append(change_indices[1:], len(arr))

            for seg_i, idx in enumerate(change_indices):
                t_left = time_vec[idx]
                for ax in axes_to_draw:
                    ax.axvline(t_left, color=color, linestyle=':', alpha=0.7)
                    
                label = str(arr[idx])
                if label and label.lower() not in ["", "sil", "<sil>", "eps", "nan"]:
                    end_idx = int(segment_ends[seg_i])
                    if end_idx < len(time_vec):
                        t_right = time_vec[end_idx]
                    else:
                        t_right = view_end
                    t_text = (t_left + t_right) * 0.5
                    self.ax_wave.text(t_text, y_pos_wave, label, color=label_color, 
                                      transform=self.ax_wave.get_xaxis_transform(),
                                      ha='center', va='bottom', fontsize=12, rotation=0)

        self.ax_wave.set_xlim(view_start, view_end)
        self.ax_param.set_xlim(view_start, view_end)
        self.set_theme(self.is_dark)
        self.canvas.draw()

    def _update_wave_plot(self):
        """Dynamically downsample waveform based on current zoom"""
        if self._wav_data is None or self._wav_total_samples <= 0:
            return
            
        xlim = self.ax_wave.get_xlim()
        fs = self._fs
        
        # Determine indices
        start_idx = int(max(0, xlim[0] * fs))
        end_idx = int(min(self._wav_total_samples, xlim[1] * fs))
        
        if end_idx <= start_idx:
             return
             
        # Extract chunk
        raw_chunk = self._wav_data[start_idx:end_idx]
        chunk = self._convert_to_float_mono(raw_chunk)
        if len(chunk) == 0:
            return
        
        # Dynamic downsampling
        step = 1
        while len(chunk) // step > 10000:
            step *= 2
            
        plot_y = chunk[::step]
        if len(plot_y) == 0:
            return
        plot_t = np.linspace(max(0, xlim[0]), min(self._wav_total_samples / fs, xlim[1]), len(plot_y))
        
        color = '#00aaff' if self.is_dark else '#007acc'
        
        # Update lines on ax_wave
        # We only touch the first line which is the waveform
        # Be careful not to delete text or vlines
        lines = [l for l in self.ax_wave.lines if l.get_linestyle() == '-' and l.get_marker() == 'None' and l.get_linewidth() == 0.5]
        
        if lines:
             lines[0].set_data(plot_t, plot_y)
        else:
             # Create line if not exists (should be created in _plot but cleared)
             self.ax_wave.plot(plot_t, plot_y, color=color, linewidth=0.5)

        max_val = float(np.max(np.abs(plot_y))) if len(plot_y) > 0 else 1.0
        if max_val == 0:
            max_val = 1.0
        self.ax_wave.set_ylim(-max_val * 1.1, max_val * 1.1)
             
        # Restore limits (set_data doesn't change limits, but just in case)
        self.ax_wave.set_xlim(xlim)
        
        self.canvas.draw_idle()

    def _convert_to_float_mono(self, data: np.ndarray) -> np.ndarray:
        arr = np.asarray(data)
        if arr.ndim > 1:
            arr = np.mean(arr, axis=1)
        if arr.dtype == np.int16:
            return arr.astype(np.float64) / 32768.0
        if arr.dtype == np.int32:
            return arr.astype(np.float64) / 2147483648.0
        if arr.dtype == np.uint8:
            return (arr.astype(np.float64) - 128.0) / 128.0
        return arr.astype(np.float64)

    def _on_scroll(self, event):
        if event.inaxes not in [self.ax_wave, self.ax_param]:
            return
        if self._wav_data is None or event.xdata is None:
            return

        cur_range = self._view_width_sec
        xdata = float(event.xdata)
        scale_factor = 0.8 if event.button == 'up' else 1.25
        new_range = cur_range * scale_factor

        total_dur = self._get_total_duration()
        max_range = min(self._max_window_sec, total_dur) if total_dur > 0 else self._max_window_sec
        if max_range <= 0:
            return

        new_range = max(0.05, min(new_range, max_range))
        rel_pos = (xdata - self._view_start_sec) / cur_range if cur_range > 0 else 0.5
        rel_pos = min(max(rel_pos, 0.0), 1.0)
        new_left = xdata - new_range * rel_pos
        self._set_view_window(new_left, new_range, redraw=True, update_slider=True)

    def _on_press(self, event):
        if event.inaxes not in [self.ax_wave, self.ax_param]: return
        if event.button == 1: # Left click
            self._pan_active = True
            self._press_x = event.xdata

    def _on_release(self, event):
        was_pan_active = self._pan_active
        self._pan_active = False
        self._press_x = None
        if was_pan_active and self._wav_data is not None:
            self._plot()

    def _on_motion(self, event):
        if not self._pan_active or self._press_x is None:
            return
        if event.inaxes not in [self.ax_wave, self.ax_param] or event.xdata is None:
            return
        if self._wav_data is None:
            return

        dx = float(event.xdata) - self._press_x
        new_left = self._view_start_sec - dx
        self._set_view_window(new_left, self._view_width_sec, redraw=False, update_slider=True)
        self._press_x = float(event.xdata)

    def _play_audio(self):
        if self._current_wav_path:
            start_ms = 0
            end_ms = 0
            if self._wav_total_samples > 0 and self._fs > 0:
                x0, x1 = self.ax_wave.get_xlim()
                total_dur = self._get_total_duration()
                left = max(0.0, min(float(x0), total_dur))
                right = max(0.0, min(float(x1), total_dur))
                if right < left:
                    left, right = right, left
                start_ms = int(left * 1000.0)
                end_ms = int(right * 1000.0)
                if end_ms <= start_ms:
                    end_ms = int(total_dur * 1000.0)

            self._pending_play_start_ms = start_ms
            self._pending_play_end_ms = end_ms
            self._playback_end_ms = None
            self._player.stop()
            self._player.setSource(QUrl.fromLocalFile(str(self._current_wav_path)))

    def _on_player_media_status_changed(self, status):
        if self._pending_play_start_ms is None or self._pending_play_end_ms is None:
            return
        if status == QMediaPlayer.MediaStatus.LoadedMedia:
            self._player.setPosition(self._pending_play_start_ms)
            self._playback_end_ms = self._pending_play_end_ms
            self._pending_play_start_ms = None
            self._pending_play_end_ms = None
            self._player.play()

    def _on_player_position_changed(self, position_ms: int):
        if self._playback_end_ms is None:
            return
        if position_ms >= self._playback_end_ms:
            self._player.stop()
            self._playback_end_ms = None

    def _save_image(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "保存图片", "plot.png", "PNG Files (*.png)")
        if path:
            self.figure.set_facecolor('white')
            self.ax_wave.set_facecolor('white')
            self.ax_param.set_facecolor('white')
            for ax in [self.ax_wave, self.ax_param]:
                ax.tick_params(colors='black', which='both')
                ax.xaxis.label.set_color('black')
                ax.yaxis.label.set_color('black')
                for spine in ax.spines.values():
                    spine.set_color('black')
                for text in ax.texts:
                    text.set_color('black')
            if hasattr(self, 'ax_param2'):
                self.ax_param2.tick_params(colors='black')
                self.ax_param2.xaxis.label.set_color('black')
                self.ax_param2.yaxis.label.set_color('black')
                for spine in self.ax_param2.spines.values():
                    spine.set_color('black')
            self.figure.savefig(path, dpi=300)
            self._plot()

    def _get_total_duration(self) -> float:
        if self._wav_total_samples <= 0 or self._fs <= 0:
            return 0.0
        return self._wav_total_samples / float(self._fs)

    def _configure_view_for_current_audio(self):
        total_duration = self._get_total_duration()
        if total_duration <= 0:
            self._view_start_sec = 0.0
            self._view_width_sec = self._max_window_sec
            self._slider_updating = True
            self.slider_position.setRange(0, 0)
            self.slider_position.setValue(0)
            self._slider_updating = False
            self.slider_position.setEnabled(False)
            self.lbl_window_range.setText("区间: -")
            return

        self._view_width_sec = min(self._max_window_sec, total_duration)
        self._view_start_sec = 0.0
        self._update_position_slider()
        self._update_window_label()

    def _normalize_view_window(self):
        total_duration = self._get_total_duration()
        if total_duration <= 0:
            self._view_start_sec = 0.0
            self._view_width_sec = self._max_window_sec
            return
        max_width = min(self._max_window_sec, total_duration)
        self._view_width_sec = min(max(self._view_width_sec, 0.05), max_width)
        max_start = max(0.0, total_duration - self._view_width_sec)
        self._view_start_sec = min(max(self._view_start_sec, 0.0), max_start)
        self._update_position_slider()
        self._update_window_label()

    def _update_position_slider(self):
        total_duration = self._get_total_duration()
        if total_duration <= 0:
            self._slider_updating = True
            self.slider_position.setRange(0, 0)
            self.slider_position.setValue(0)
            self._slider_updating = False
            self.slider_position.setEnabled(False)
            return

        max_start = max(0.0, total_duration - self._view_width_sec)
        slider_max = int(round(max_start * 1000.0))
        slider_value = int(round(self._view_start_sec * 1000.0))
        slider_value = min(max(slider_value, 0), slider_max)
        self._slider_updating = True
        self.slider_position.setRange(0, slider_max)
        self.slider_position.setValue(slider_value)
        self._slider_updating = False
        self.slider_position.setEnabled(slider_max > 0)

    def _update_window_label(self):
        total_duration = self._get_total_duration()
        if total_duration <= 0:
            self.lbl_window_range.setText("区间: -")
            return
        start = self._view_start_sec
        end = self._view_start_sec + self._view_width_sec
        self.lbl_window_range.setText(f"区间: {start:.2f}-{end:.2f}s / {total_duration:.2f}s")

    def _set_view_window(self, start_sec: float, width_sec: Optional[float] = None, redraw: bool = True, update_slider: bool = True):
        total_duration = self._get_total_duration()
        if total_duration <= 0:
            return
        max_width = min(self._max_window_sec, total_duration)
        if width_sec is None:
            width = self._view_width_sec
        else:
            width = width_sec
        width = max(0.05, min(width, max_width))
        max_start = max(0.0, total_duration - width)
        start = min(max(0.0, start_sec), max_start)
        self._view_start_sec = start
        self._view_width_sec = width
        self._update_window_label()
        if update_slider:
            self._update_position_slider()
        self.ax_wave.set_xlim(self._view_start_sec, self._view_start_sec + self._view_width_sec)
        self.ax_param.set_xlim(self._view_start_sec, self._view_start_sec + self._view_width_sec)
        if redraw:
            self._plot()
        else:
            self._update_wave_plot()

    def _on_position_slider_changed(self, value: int):
        if self._slider_updating:
            return
        start_sec = float(value) / 1000.0
        self._set_view_window(start_sec, self._view_width_sec, redraw=True, update_slider=False)

    def set_theme(self, is_dark: bool):
        self.is_dark = is_dark
        if is_dark:
            facecolor = '#1e1e1e'
            axis_color = 'white'
        else:
            facecolor = '#f0f0f0'
            axis_color = 'black'
            
        self.figure.set_facecolor(facecolor)
        self.ax_wave.set_facecolor(facecolor)
        self.ax_param.set_facecolor(facecolor)
        
        for ax in [self.ax_wave, self.ax_param]:
            ax.tick_params(colors=axis_color, which='both')
            if ax == self.ax_param:
                ax.tick_params(labelbottom=True, bottom=True)
            elif ax == self.ax_wave:
                ax.tick_params(labelbottom=False, bottom=False)
            ax.xaxis.label.set_color(axis_color)
            ax.yaxis.label.set_color(axis_color)
            for spine in ax.spines.values():
                spine.set_color(axis_color)
                
        if hasattr(self, 'ax_param2'):
            self.ax_param2.tick_params(colors=axis_color)
            for spine in self.ax_param2.spines.values():
                spine.set_color(axis_color)

        self.canvas.draw()

    def _show_param_help(self):
        dialog = SharedParameterHelpDialog(self)
        dialog.exec()

    def _open_help_doc(self):
        help_file = Path(r"d:\PhoneticToolbox\PhoneticToolbox_v2\Phonetic_Export\index.html")
        if not help_file.exists():
            QtWidgets.QMessageBox.warning(self, "帮助", f"未找到帮助文件：{help_file}")
            return
        url = QUrl.fromLocalFile(str(help_file))
        url.setFragment("s1764337355643")
        opened = QtGui.QDesktopServices.openUrl(url)
        if not opened:
            QtWidgets.QMessageBox.warning(self, "帮助", "帮助页面打开失败。")
